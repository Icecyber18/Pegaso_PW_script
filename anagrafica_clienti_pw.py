# script creato da Pastore Pietro come richiesto nel PW Pegaso
# script per generare un file Excel con un'anagrafica di clienti
# con nome, cognome, email, telefono e due hash delle password
# SHA256 e SHA512

# librerie necessarie
import random          # Serve per generare numeri e scelte casuali
import string         # Serve per utilizzare lettere e caratteri speciali
import hashlib        # Serve per crittografare le password
import os             # Serve per operazioni sul sistema operativo (creazione cartelle)
from openpyxl import Workbook    # Serve per creare e gestire file Excel
import faker          # Serve per generare vari tipi di dati casuali realistici, come nomi, cognomi, email, telefono, etc.
                      # ho preferito utilizzare faker una libreria più avanzata rispetto a names e genera dati più realistici

def genera_password():
    # per generarepassword casuale di 12 caratteri
    caratteri = string.ascii_letters + string.digits + string.punctuation
    password = ''.join(random.choice(caratteri) for _ in range(12))
    
    # Crea gli hash della password, ho preferito generare due hash diversi
    # per aumentare la sicurezza ed evitare ipotetiche collisioni
    hash_256 = hashlib.sha256(password.encode()).hexdigest()
    hash_512 = hashlib.sha512(password.encode()).hexdigest()
    
    return hash_256, hash_512

def genera_numero_telefono():
    # genera un numero di telefono italiano casuale
    prefissi_italiani = ['320', '330', '340', '350', '360', '370', '380', '390']
    prefisso = random.choice(prefissi_italiani)
    numero = random.randint(1000000, 9999999)
    
    return f"+39{prefisso}{numero}"

# funzione per chiedere il numero di utenti da generare, utile per generare un data set di test
def chiedi_numero_utenti():
    while True:
        try:
            numero = int(input("Inserisci il numero di utenti da generare: "))
            if numero > 0:
                return numero
            print("Per favore inserisci un numero positivo")
        except ValueError:
            print("Per favore inserisci un numero valido")

#
def crea_cartella_output():
    cartella = "C:/Project_Work_Pegaso"
    if not os.path.exists(cartella):
        os.makedirs(cartella)
    return cartella

# funzione per creare il file Excel
def crea_excel():
    workbook = Workbook()
    foglio = workbook.active
    foglio.title = "Anagrafica Clienti"
    
    # inserisco le intestazioni all'interno del file Excel
    intestazioni = ['Nome', 'Cognome', 'Email', 'Telefono', 'Password Hash (SHA256)', 'Password Hash (SHA512)']
    for colonna, intestazione in enumerate(intestazioni, 1):
        foglio.cell(row=1, column=colonna, value=intestazione)
    
    return workbook, foglio

# funzione principale del programma
def main():
    numero_utenti = chiedi_numero_utenti()
    cartella = crea_cartella_output()
    fake = faker.Faker('it_IT') 
    workbook, foglio = crea_excel()
    
    # per ogni riga del file Excel, genera i dati degli utenti
    for riga in range(2, numero_utenti + 2):
        # Genera dati utente usando faker per nomi e cognomiitaliani
        nome = fake.first_name()
        cognome = fake.last_name()
        # genera un email composta da nome.cognome@dominio.it
        email = f"{nome.lower()}.{cognome.lower()}@{fake.free_email_domain()}" 
        telefono = genera_numero_telefono()
        hash_256, hash_512 = genera_password()
        
        # Inserisce dati nel foglio Excel
        dati_riga = [nome, cognome, email, telefono, hash_256, hash_512]
        for colonna, valore in enumerate(dati_riga, 1):
            foglio.cell(row=riga, column=colonna, value=valore)
    
    # salva il file Excel
    percorso_file = os.path.join(cartella, "Anagrafica_clienti_Pw_Pegaso.xlsx")
    workbook.save(percorso_file)
    print(f"\nFile generato con successo in: {percorso_file}")


if __name__ == "__main__":
    main()  # Chiama la funzione principale
